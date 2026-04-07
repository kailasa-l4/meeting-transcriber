"""XLSX writer skill — save leads to a local Excel file."""

import os
import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from src.config import settings

logger = logging.getLogger(__name__)

# Column headers matching PRD sheet mapping
COLUMNS = [
    "Name", "Role / Title", "Company Name", "Email", "Phone",
    "WhatsApp", "Website", "Details", "Source",
    "LinkedIn", "Location", "Confidence Score", "Country",
]

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def write_leads_to_xlsx(
    leads: list[dict],
    country: str,
    job_id: str = "",
    output_dir: str | None = None,
) -> str:
    """Write leads to an Excel file.

    Creates or appends to a file named `gold_leads_{country}.xlsx`.
    Returns the file path.
    """
    out_dir = Path(output_dir or settings.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    filename = f"gold_leads_{country.lower().replace(' ', '_')}.xlsx"
    filepath = out_dir / filename

    if filepath.exists():
        wb = load_workbook(filepath)
        ws = wb.active
        logger.info(f"Appending to existing file: {filepath}")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = f"{country} Leads"
        _write_header(ws)
        logger.info(f"Creating new file: {filepath}")

    rows_written = 0
    for lead in leads:
        row = [
            lead.get("name", ""),
            lead.get("role_title", ""),
            lead.get("company_name", ""),
            lead.get("email", ""),
            lead.get("phone", ""),
            lead.get("whatsapp", ""),
            lead.get("website", ""),
            lead.get("details", ""),
            lead.get("source_text", ""),
            lead.get("linkedin_url", ""),
            lead.get("location", ""),
            lead.get("confidence_score", 0),
            country,
        ]
        ws.append(row)
        rows_written += 1

    # Auto-fit column widths (approximate)
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

    # Add metadata row
    ws.append([])
    ws.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", f"Job: {job_id}", f"Leads: {rows_written}"])

    wb.save(filepath)
    logger.info(f"Wrote {rows_written} leads to {filepath}")
    return str(filepath)


def _write_header(ws):
    """Write styled header row."""
    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
    ws.freeze_panes = "A2"


SKILL_METADATA = {
    "id": "xlsx-writer",
    "version": "1.0.0",
    "description": "Save leads to local Excel file",
}
